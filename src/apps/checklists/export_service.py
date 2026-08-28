"""Сервис для экспорта данных чек-листов во внешние форматы (Excel, PDF)."""

import os
from io import BytesIO
from xml.sax.saxutils import escape

import openpyxl
from django.conf import settings
from openpyxl.styles import Alignment, Border, Font, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.checklists.interfaces import IChecklistBuilder
from apps.checklists.models import ChecklistResult


class ChecklistExportDirector:
    """
    Директор (Оркестратор).

    Определить строгий порядок вызова шагов строительства документа.
    """

    def __init__(self, builder: IChecklistBuilder):
        self._builder = builder

    def construct_document(self) -> bytes:
        """
        Запустить процесс строительства документа шаг за шагом.

        Вернуть готовый файл в байтах.
        """
        self._builder.build_header()
        self._builder.build_body()
        self._builder.build_footer()
        return self._builder.get_document_bytes()


class ExcelChecklistBuilder(IChecklistBuilder):
    """
    Сервис для генерации печатных форм анкет в формате Microsoft Excel (.xlsx).

    Следовать принципу единой ответственности (SRP): инкапсулировать логику работы
    с библиотекой openpyxl, стилизацию ячеек и компоновку данных.
    """

    def __init__(self, result: ChecklistResult):
        self.result = result
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Журнал смены"
        self.ws.sheet_view.showGridLines = False

        self.current_row = 1

        self.font_bold = Font(bold=True, size=11)
        self.font_normal = Font(size=11)
        self.font_title = Font(size=10)
        self.font_equipment = Font(bold=True, size=14)

        self.align_center = Alignment(horizontal='center', vertical='center')
        self.align_right = Alignment(horizontal='right', vertical='center')
        self.align_left = Alignment(horizontal='left', vertical='center')
        self.align_wrap = Alignment(horizontal='left', vertical='center',
                                    wrap_text=True)

        self.border_thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    def build_header(self) -> None:
        """Сгенерировать шапку документа с метаданными (Дата, Смена, Время)."""
        self.ws.merge_cells('A1:E1')
        self.ws.cell(row=1, column=1,
                     value="Приложение № 1").alignment = self.align_center

        self.ws.merge_cells('A2:E2')
        self.ws.cell(row=2, column=1,
                     value="Журнал приема-передачи смены").alignment = self.align_center

        self.ws.cell(row=3, column=5,
                     value="Код документа ФЗ-Ж04-П1").alignment = self.align_right
        self.current_row = 5

        date_str = self.result.created_at.strftime("%d.%m.%Y")
        shift_str = self.result.get_shift_number_display() or "___"
        time_str = self.result.shift_time.strftime(
            "%H:%M") if self.result.shift_time else "___"

        self.ws.cell(row=self.current_row, column=1,
                     value=f"Дата  {date_str}").font = self.font_normal
        self.ws.merge_cells(start_row=self.current_row, start_column=2,
                            end_row=self.current_row, end_column=3)
        self.ws.cell(row=self.current_row, column=2,
                     value=f"Смена(1-дневная, 2-ночная)  {shift_str}")
        self.ws.cell(row=self.current_row, column=5,
                     value=f"Время смены  {time_str}").alignment = self.align_right

        self.current_row += 2

    def build_body(self) -> None:
        """
        Сгенерировать основное тело анкеты.

        Выводит название оборудования, группирует вопросы и записывает
        ответы пользователя с комментариями.
        """
        self.ws.cell(row=self.current_row, column=1,
                     value="Техническое состояние оборудования").font = self.font_bold
        self.ws.cell(row=self.current_row, column=2,
                     value=self.result.template.equipment_uid).font = self.font_equipment
        self.current_row += 1

        answers = self.result.answers.select_related('field__group').order_by(
            'field__group__order', 'field__order')
        current_group = None

        for ans in answers:
            group_name = ans.field.group.name if ans.field.group else "Дополнительно"

            if current_group != group_name:
                current_group = group_name
                self.ws.merge_cells(start_row=self.current_row, start_column=1,
                                    end_row=self.current_row, end_column=5)
                self.ws.cell(row=self.current_row, column=1,
                             value=group_name).font = self.font_bold
                self._apply_row_borders(self.current_row)
                self.current_row += 1

            opt1, opt2 = self._get_options_display(ans)
            comment_val = ans.comment if str(ans.comment).strip() else ""

            self.ws.cell(row=self.current_row, column=1, value=ans.field.name)
            self.ws.cell(row=self.current_row, column=2, value=opt1)
            self.ws.cell(row=self.current_row, column=3, value=opt2)
            self.ws.cell(row=self.current_row, column=4, value="Замечания")
            self.ws.cell(row=self.current_row, column=5, value=comment_val)

            self._apply_row_borders(self.current_row)
            self.current_row += 1

    def build_footer(self) -> None:
        """
        Сгенерировать нижнюю часть документа.

        Отображает общий комментарий к смене и подписи ответственных лиц.
        """
        sigs = {sig.role: sig for sig in self.result.signatures.all()}
        operator_out = sigs.get('AUTHOR')
        operator_in = sigs.get('READER')
        master = sigs.get('APPROVER')

        self.ws.merge_cells(start_row=self.current_row, start_column=1,
                            end_row=self.current_row, end_column=3)
        self.ws.cell(row=self.current_row, column=1,
                     value=f"Смену сдал    {operator_out.user_uid if operator_out else ''}")
        self.ws.cell(row=self.current_row, column=4, value="Замечания")
        self.ws.cell(row=self.current_row, column=5,
                     value=self.result.general_comment)
        self._apply_row_borders(self.current_row)
        self.current_row += 1

        self.ws.merge_cells(start_row=self.current_row, start_column=1,
                            end_row=self.current_row, end_column=3)
        self.ws.cell(row=self.current_row, column=1,
                     value=f"Смену принял  {operator_in.user_uid if operator_in else ''}")
        self.ws.cell(row=self.current_row, column=4, value="Замечания")
        self.ws.cell(row=self.current_row, column=5, value="")
        self._apply_row_borders(self.current_row)
        self.current_row += 2

        master_uid = master.user_uid if master else "________________________"
        self.ws.merge_cells(start_row=self.current_row, start_column=1,
                            end_row=self.current_row, end_column=5)
        self.ws.cell(row=self.current_row, column=1,
                     value=f"Мастер смены  {master_uid}").alignment = self.align_center

    def get_document_bytes(self) -> bytes:
        """Настроить ширину колонок и выгрузить Excel-файл в байты."""
        self.ws.column_dimensions['A'].width = 45
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 15
        self.ws.column_dimensions['D'].width = 12
        self.ws.column_dimensions['E'].width = 45

        stream = BytesIO()
        self.wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    def _apply_row_borders(self, row, start_col=1, end_col=7):
        """Вспомогательный метод для отрисовки сетки на всю строку."""
        for col in range(start_col, end_col + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.border = self.border_thin
            cell.alignment = self.align_wrap

    def _get_options_display(self, ans):
        """Формирует две колонки с вариантами ответов и метками [V] / [ ]"""
        opt1, opt2 = "", ""
        val = str(ans.value).strip()

        if ans.field.field_type in ['CHECKBOX', 'RADIO']:
            is_true = val.lower() == 'true'
            opt1 = "[V] Да" if is_true else "[ ] Да"
            opt2 = "[V] Нет" if not is_true and val else "[ ] Нет"
        elif ans.field.field_type == 'CHOICE':
            choices = list(ans.field.choices.all())
            if len(choices) >= 2:
                c1, c2 = choices[0].value, choices[1].value
                opt1 = f"[V] {c1}" if val == c1 else f"[ ] {c1}"
                opt2 = f"[V] {c2}" if val == c2 else f"[ ] {c2}"
            elif len(choices) == 1:
                c1 = choices[0].value
                opt1 = f"[V] {c1}" if val == c1 else f"[ ] {c1}"
            else:
                opt1 = val
        else:
            opt1 = val

        return opt1, opt2


class PdfChecklistBuilder(IChecklistBuilder):
    """
    Сервис для экспорта анкеты в PDF-формат.

    Сгенерировать документ по структуре, аналогичной Excel (печатный бланк).
    """

    def __init__(self, result: ChecklistResult):
        self.result = result
        self.elements = []

        font_path = os.path.join(settings.BASE_DIR, 'arial.ttf')
        if not os.path.exists(font_path):
            raise FileNotFoundError(
                f"ОШИБКА: Файл шрифта не найден по пути {font_path}! "
                "Чтобы PDF работал с русским языком, скачайте файл arial.ttf "
                "и положите его в корень проекта (рядом с manage.py)."
            )

        pdfmetrics.registerFont(TTFont('Arial', font_path))
        self.font_name = 'Arial'

        styles = getSampleStyleSheet()
        self.p_normal = ParagraphStyle(
            'CellNormal', parent=styles['Normal'], fontName=self.font_name,
            fontSize=9, leading=11
        )
        self.p_violation = ParagraphStyle(
            'CellViolation', parent=styles['Normal'], fontName=self.font_name,
            fontSize=9, leading=11, textColor=colors.red
        )
        self.p_group = ParagraphStyle(
            'CellGroup', parent=styles['Normal'], fontName=self.font_name,
            fontSize=10, leading=12, fontBold=True
        )

    def build_header(self) -> None:
        """
        Сформировать верхний блок PDF-документа.

        Отобразить дату создания, тип смены и время смены в одну строку.
        """
        styles = getSampleStyleSheet()
        center_style = ParagraphStyle('Center', parent=styles['Normal'],
                                      fontName=self.font_name, alignment=1)
        right_style = ParagraphStyle('Right', parent=styles['Normal'],
                                     fontName=self.font_name, alignment=2)

        self.elements.append(Paragraph("Приложение № 1", center_style))
        self.elements.append(
            Paragraph("Журнал приема-передачи смены", center_style))
        self.elements.append(Paragraph("Код документа ФЗ-Ж04-П1", right_style))
        self.elements.append(Spacer(1, 15))

        date_str = self.result.created_at.strftime("%d.%m.%Y")
        shift_str = self.result.get_shift_number_display() or "___"
        time_str = self.result.shift_time.strftime(
            "%H:%M") if self.result.shift_time else "___"

        data = [
            [f"Дата  {date_str}", f"Смена(1-дневная, 2-ночная)  {shift_str}",
             f"Время смены  {time_str}"]]
        t = Table(data, colWidths=[150, 250, 140])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        self.elements.append(t)
        self.elements.append(Spacer(1, 10))

    def build_body(self) -> None:
        """
        Сформировать основное тело PDF-документа.

        Вывести UID оборудования, а затем построить таблицу со списком вопросов,
        сгруппированных по смысловым блокам, с ответами и подчеркнутыми комментариями.
        Отклонения автоматически подсвечиваются красным.
        """
        styles = getSampleStyleSheet()
        bold_style = ParagraphStyle('Bold', parent=styles['Normal'],
                                    fontName=self.font_name, fontBold=True)
        eq_style = ParagraphStyle('Eq', parent=styles['Normal'],
                                  fontName=self.font_name, fontSize=12,
                                  fontBold=True)

        data_head = [
            [Paragraph("<b>Техническое состояние оборудования</b>",
                       bold_style),
             Paragraph(f"<b>{self.result.template.equipment_uid}</b>",
                       eq_style)]
        ]
        th = Table(data_head, colWidths=[250, 290])
        self.elements.append(th)

        data = []
        table_styles = [
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]

        row_idx = 0
        answers = self.result.answers.select_related('field__group').order_by(
            'field__group__order', 'field__order')
        current_group = None

        for ans in answers:
            group_name = ans.field.group.name if ans.field.group else "Дополнительно"

            if current_group != group_name:
                current_group = group_name
                data.append(
                    [Paragraph(f"<b>{escape(group_name)}</b>", self.p_group),
                     "", "", "", ""])
                table_styles.append(('SPAN', (0, row_idx), (4, row_idx)))
                table_styles.append(('BACKGROUND', (0, row_idx), (4, row_idx),
                                     colors.lightgrey))
                row_idx += 1

            opt1, opt2 = self._get_options_display(ans)
            comment_val = ans.comment if str(ans.comment).strip() else ""

            opt_style = self.p_violation if ans.is_violation else self.p_normal

            data.append([
                Paragraph(escape(ans.field.name), self.p_normal),
                Paragraph(escape(opt1), opt_style),
                Paragraph(escape(opt2), opt_style),
                Paragraph("Замечания", self.p_normal),
                Paragraph(escape(comment_val), self.p_normal)
            ])
            row_idx += 1

        t = Table(data, colWidths=[200, 70, 70, 60, 140])
        t.setStyle(TableStyle(table_styles))
        self.elements.append(t)

    def build_footer(self) -> None:
        """
        Сформировать нижний блок PDF-документа.

        Отрисовать поля для подписей ответственных лиц и поле для вывода
        общего комментария по смене.
        """
        sigs = {sig.role: sig for sig in self.result.signatures.all()}
        operator_out = sigs.get('AUTHOR')
        operator_in = sigs.get('READER')
        master = sigs.get('APPROVER')

        data = []
        table_styles = [
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]

        out_uid = operator_out.user_uid if operator_out else ""
        data.append([
            Paragraph(f"<b>Смену сдал</b> {escape(out_uid)}", self.p_normal),
            "", "",
            Paragraph("Замечания", self.p_normal),
            Paragraph(escape(self.result.general_comment or ""), self.p_normal)
        ])
        table_styles.append(('SPAN', (0, 0), (2, 0)))

        in_uid = operator_in.user_uid if operator_in else ""
        data.append([
            Paragraph(f"<b>Смену принял</b> {escape(in_uid)}", self.p_normal),
            "", "", "", ""
        ])
        table_styles.append(('SPAN', (0, 1), (2, 1)))

        t = Table(data, colWidths=[113, 113, 114, 60, 140])
        t.setStyle(TableStyle(table_styles))
        self.elements.append(t)
        self.elements.append(Spacer(1, 20))

        styles = getSampleStyleSheet()
        center_style = ParagraphStyle('Center', parent=styles['Normal'],
                                      fontName=self.font_name, alignment=1)
        master_uid = master.user_uid if master else "________________________"
        self.elements.append(
            Paragraph(f"Мастер смены {master_uid}", center_style))

    def get_document_bytes(self) -> bytes:
        """
        Скомпилировать все собранные элементы в итоговый документ.

        Вернуть байт-строку файла формата A4.
        """
        stream = BytesIO()
        doc = SimpleDocTemplate(stream, pagesize=A4, rightMargin=25,
                                leftMargin=30, topMargin=30, bottomMargin=30)
        doc.build(self.elements)
        return stream.getvalue()

    def _get_options_display(self, ans):
        """Такая же логика меток [V], как и для Excel"""
        opt1, opt2 = "", ""
        val = str(ans.value).strip()

        if ans.field.field_type in ['CHECKBOX', 'RADIO']:
            is_true = val.lower() == 'true'
            opt1 = "[V] Да" if is_true else "[ ] Да"
            opt2 = "[V] Нет" if not is_true and val else "[ ] Нет"
        elif ans.field.field_type == 'CHOICE':
            choices = list(ans.field.choices.all())
            if len(choices) >= 2:
                c1, c2 = choices[0].value, choices[1].value
                opt1 = f"[V] {c1}" if val == c1 else f"[ ] {c1}"
                opt2 = f"[V] {c2}" if val == c2 else f"[ ] {c2}"
            elif len(choices) == 1:
                c1 = choices[0].value
                opt1 = f"[V] {c1}" if val == c1 else f"[ ] {c1}"
            else:
                opt1 = val
        else:
            opt1 = val

        return opt1, opt2
